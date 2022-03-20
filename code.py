import re
import os
import openpyxl
from pathlib import Path

def sorted_alphanumeric(data):
    convert = lambda text: int(text) if text.isdigit() else text.lower()
    alphanum_key = lambda key: [ convert(c) for c in re.split('([0-9]+)', key) ]
    return sorted(data, key=alphanum_key)

def list_files(my_directory, output_name):
    my_path = Path(my_directory)
    list_files = sorted_alphanumeric(os.listdir(my_path))
    my_len_list = len(list_files)
    file_interest_list = []
    for i in range(0,my_len_list):
        file_path = my_path/list_files[i]
        file_interest_list = file_interest_list + [file_path.stem]
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    length_list = len(file_interest_list)
    sheet.cell(row=1, column=1).value = 'S/N'
    sheet.cell(row=1, column=2).value = 'File Name'
    for i in range(1, length_list + 1):
        sheet.cell(row=i+1, column=1).value = i
        sheet.cell(row=i+1, column=2).value = file_interest_list[i-1]
    wb.save(output_name)
